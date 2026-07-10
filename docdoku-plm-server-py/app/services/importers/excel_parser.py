"""Excel 属性解析器——对齐 Payara ExcelParser.java。

解析 .xlsx 文件，提取零件属性定义和数据行，做校验并收集错误。
纯逻辑，不碰数据库。
"""

import datetime
import io
import re
import urllib.parse
from dataclasses import dataclass, field

from openpyxl import load_workbook

# 合法类型集合（规范化后）
_VALID_TYPES: set[str] = {"TEXT", "NUMBER", "DATE", "BOOLEAN", "URL", "LONG_TEXT", "LOV"}

# 表头正则（对齐 Java ExcelParser.java）
# 先试 LOV 再试 ATT；使用 fullmatch（Java matches() 即全匹配）
_PATTERN_NEW_LOV = re.compile(r"(.*) <(.*)> <(.*)>")
_PATTERN_NEW_ATT = re.compile(r"(.*) <(.*)>")

# 属性 ID 必须匹配纯数字（空串代表"新建"，也通过）
_ID_PATTERN = re.compile(r"^[0-9]*$")


@dataclass
class ParsedAttribute:
    """解析后的单个属性。"""
    name: str
    type: str                     # 规范 token: TEXT/NUMBER/DATE/BOOLEAN/URL/LONG_TEXT/LOV
    value: str | None             # 原始字符串值（未做类型转换）
    attribute_id: int | None = None   # 数据单元格 comment 里的属性 ID；None=新建
    lov_name: str | None = None   # 仅 LOV 有


@dataclass
class ParsedPart:
    """解析后的单个零件（parts 导入）。"""
    number: str
    attributes: list[ParsedAttribute] = field(default_factory=list)
    # pathdata 导入额外携带定位信息（best-effort）
    product_id: str | None = None
    serial_number: str | None = None


@dataclass
class ExcelParseResult:
    """Excel 解析完整结果。"""
    import_type: str              # "parts" | "pathdata"
    parts: list[ParsedPart] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _normalize_type(type_str: str | None) -> str | None:
    """将捕获到的类型字符串转为规范 token（大写）。"""
    if not type_str:
        return None
    upper = type_str.strip().upper()
    if upper == "TEXT":
        return "TEXT"
    if upper == "NUMBER":
        return "NUMBER"
    if upper == "DATE":
        return "DATE"
    if upper == "BOOLEAN":
        return "BOOLEAN"
    if upper == "URL":
        return "URL"
    if upper in ("LONG_TEXT", "LONGTEXT"):
        return "LONG_TEXT"
    if upper == "LISTOFVALUES":
        return "LOV"
    return None


def _validate_value(value: str | None, attr_type: str, attr_name: str,
                    row: int, errors: list[str]):
    """校验单个属性值（非空时做类型校验，空值时按类型决定是否允许）。"""
    if value is None or str(value).strip() == "":
        # 空值处理：BOOLEAN/NUMBER/LOV 不允许空值
        if attr_type in ("BOOLEAN", "NUMBER", "LOV"):
            errors.append(
                f"EMPTY_FIELD: column '{attr_name}' row {row}: "
                f"attribute type '{attr_type}' does not allow empty value"
            )
        # TEXT/DATE/URL/LONG_TEXT 允许空 → 不做校验
        return

    val = str(value)

    if attr_type == "TEXT":
        if len(val) > 255:
            errors.append(
                f"INVALID_TEXT_VALUE: column '{attr_name}' row {row}: "
                f"text exceeds 255 characters (length={len(val)})"
            )

    elif attr_type == "BOOLEAN":
        if val not in ("true", "false"):
            errors.append(
                f"INVALID_BOOLEAN_VALUE: column '{attr_name}' row {row}: '{val}'"
            )

    elif attr_type == "DATE":
        try:
            datetime.datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            errors.append(
                f"INVALID_DATE_VALUE: column '{attr_name}' row {row}: '{val}'"
            )

    elif attr_type == "NUMBER":
        try:
            float(val)
        except (ValueError, TypeError):
            errors.append(
                f"INVALID_NUMBER_VALUE: column '{attr_name}' row {row}: '{val}'"
            )

    elif attr_type == "URL":
        parsed = urllib.parse.urlparse(val)
        if parsed.scheme not in ("http", "https", "ftp") or not parsed.netloc:
            errors.append(
                f"INVALID_URL_VALUE: column '{attr_name}' row {row}: '{val}'"
            )

    elif attr_type == "LOV":
        # LOV 值不在此处校验（留待写库时查 lov 表）
        pass


def _parse_header_cell(cell_value, cell_comment, col_index: int,
                       errors: list[str]) -> tuple[str | None, str | None, str | None]:
    """解析单个表头单元格，返回 (attribute_name, normalized_type, lov_name)。

    返回 None 表示该列不参与属性解析（如 pm.number 列）。
    """
    header_text = str(cell_value).strip() if cell_value is not None else ""

    # 1. 尝试 LOV 正则
    m = _PATTERN_NEW_LOV.fullmatch(header_text)
    if m:
        name = m.group(1).strip()
        type_candidate = m.group(2).strip()
        lov_name_val = m.group(3).strip()
        if type_candidate == "ListOfValues":
            return name, "LOV", lov_name_val
        # 即使 group(2) 不是 ListOfValues，也当成 LOV？
        # Java 逻辑：if (m.matches() && "ListOfValues".equals(m.group(2))) → LOV
        # else → 不匹配，继续往下走。所以这里要 fall through。
        # 实际上 LOV 正则有 3 个 group，如果 group(2) != "ListOfValues"，不应作为 LOV。
        # 但两个正则中 LOV 的正则更贪婪，如果它匹配了，ATT 也一定匹配。
        # 所以 LOV match 但 group(2) != "ListOfValues" 时，应该让 ATT 来处理。
        pass

    # 2. 尝试普通属性正则
    m = _PATTERN_NEW_ATT.fullmatch(header_text)
    if m:
        name = m.group(1).strip()
        type_str = m.group(2).strip()
        norm_type = _normalize_type(type_str)
        if norm_type is None:
            errors.append(
                f"ATTRIBUTE_TYPE_NOT_FOUND: column {col_index}: "
                f"unknown type '{type_str}' for attribute '{name}'"
            )
        return name, norm_type, None

    # 3. 两个正则都不命中 → 必须带 cell comment
    if cell_comment is None:
        errors.append(
            f"MISSING_COMMENT: column {col_index}: "
            f"header '{header_text}' has no type pattern and no comment"
        )
        return None, None, None

    type_str = str(cell_comment).strip()
    norm_type = _normalize_type(type_str)
    if norm_type is None:
        errors.append(
            f"ATTRIBUTE_TYPE_NOT_FOUND: column {col_index}: "
            f"unknown type '{type_str}' for header '{header_text}'"
        )
        return None, None, None

    return header_text, norm_type, None


def _parse_ids_from_comment(cell_comment, col_name: str, row: int,
                            errors: list[str]) -> list[str] | None:
    """从单元格批注中解析属性 ID 列表。

    返回 None 表示该单元格无批注（全部视为新建，不触发 MISSING_ATTRIBUTE_ID）。
    返回 list 表示有批注，按 | 拆分，每个元素为 ID 字符串。
    """
    if cell_comment is None:
        return None

    comment_text = str(cell_comment).strip()
    if comment_text == "":
        return None

    ids = comment_text.split("|")
    for id_str in ids:
        if not _ID_PATTERN.match(id_str.strip()):
            errors.append(
                f"INVALID_ATTRIBUTE_ID: column '{col_name}' row {row}: "
                f"'{id_str}' is not a valid attribute ID"
            )
    return ids


def _parse_data_cell(value, comment, col_def: dict, row: int,
                     errors: list[str]) -> list[ParsedAttribute]:
    """解析单个数据单元格，返回 ParsedAttribute 列表（多值拆分）。"""
    attr_name = col_def["name"]
    attr_type = col_def["type"]
    lov_name = col_def.get("lov_name")

    if value is None:
        # 空值 → 生成单个属性（空值校验在 _validate_value 中完成）
        result = ParsedAttribute(
            name=attr_name,
            type=attr_type,
            value=None,
            attribute_id=None,
            lov_name=lov_name,
        )
        _validate_value(None, attr_type, attr_name, row, errors)
        return [result]

    val_str = str(value)
    values = [v.strip() for v in val_str.split("|")]

    # 解析 ID（从 comment）；None 表示无批注 → 全部视为新建
    ids = _parse_ids_from_comment(comment, attr_name, row, errors)

    if ids is not None and len(values) > len(ids):
        errors.append(
            f"MISSING_ATTRIBUTE_ID: column '{attr_name}' row {row}: "
            f"have {len(values)} value(s) but only {len(ids)} attribute ID(s)"
        )

    result = []
    for i, val in enumerate(values):
        # 确定 attribute_id
        attr_id: int | None = None
        if ids is not None and i < len(ids):
            id_str = ids[i].strip()
            # 只对纯数字（含空串）的 ID 做转换；非纯数字已在 _parse_ids_from_comment 报错
            if id_str != "" and _ID_PATTERN.match(id_str):
                attr_id = int(id_str)
        # ids 不够、ids=None、或 id 为空串 → attribute_id=None（新建）

        # 值校验
        _validate_value(val if val != "" else None, attr_type, attr_name, row, errors)

        result.append(ParsedAttribute(
            name=attr_name,
            type=attr_type,
            value=val if val != "" else None,
            attribute_id=attr_id,
            lov_name=lov_name,
        ))

    return result


def parse_excel(data: bytes, import_type: str = "parts") -> ExcelParseResult:
    """解析 .xlsx 字节流。校验+解析一次完成：先做表头/首列校验与类型校验，
    错误收集进 result.errors；即使有错误也尽量返回已解析的 parts（供 dry-run 展示）。
    """
    result = ExcelParseResult(import_type=import_type)

    # 加载工作簿
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        result.errors.append("EMPTY_FILE: cannot open file, invalid xlsx format")
        return result

    if not wb.worksheets:
        result.errors.append("EMPTY_FILE: workbook has no sheets")
        return result

    ws = wb.worksheets[0]

    if ws.max_row is None or ws.max_row < 1:
        result.errors.append("EMPTY_FILE: empty sheet")
        return result

    # 读取总列数（第一行）
    max_col = ws.max_column or 0

    # 空 sheet / 无列 → EMPTY_FILE
    # 注意：新 Workbook 默认 sheet 的 dimension 可能为 A1:A1，需额外检查 A1 是否有内容
    if max_col == 0:
        result.errors.append("EMPTY_FILE: sheet has no columns")
        return result
    if max_col <= 1:
        a1 = ws.cell(1, 1)
        if (a1.value is None or str(a1.value).strip() == "") and a1.comment is None:
            result.errors.append("EMPTY_FILE: sheet is empty")
            return result

    # ========== 首列/首行标识校验 ==========
    if import_type == "parts":
        if max_col < 2:
            result.errors.append("INVALID_HEADER: need at least 2 columns for parts import")
            return result
        a1_comment = ws.cell(1, 1).comment
        if a1_comment is None or str(a1_comment.text).strip() != "pm.number":
            result.errors.append(
                "INVALID_HEADER: first column (A1) must have comment 'pm.number'"
            )
            # 仍然继续解析，返回部分结果

    elif import_type == "pathdata":
        if max_col <= 3:
            result.errors.append(
                "INVALID_COLUMNS_NUMBER: need >3 columns for pathdata import"
            )
            return result
        a1 = ws.cell(1, 1).comment
        b1 = ws.cell(1, 2).comment
        c1 = ws.cell(1, 3).comment
        valid = True
        if a1 is None or str(a1.text).strip() != "ctx.productId":
            result.errors.append(
                "INVALID_HEADER: first column (A1) must have comment 'ctx.productId'"
            )
            valid = False
        if b1 is None or str(b1.text).strip() != "ctx.serialNumber":
            result.errors.append(
                "INVALID_HEADER: second column (B1) must have comment 'ctx.serialNumber'"
            )
            valid = False
        if c1 is None or str(c1.text).strip() != "pm.number":
            result.errors.append(
                "INVALID_HEADER: third column (C1) must have comment 'pm.number'"
            )
            valid = False
        if not valid:
            return result

    # ========== 解析表头 ==========
    # 为每个列建立属性定义（None 表示该列不参与属性解析）
    attr_columns: list[dict | None] = []
    if import_type == "parts":
        # 第1列是 pm.number，不是属性列
        attr_columns.append(None)  # col 1 = pm.number
        attr_start = 2
    else:  # pathdata
        # 前三列是 ctx.productId, ctx.serialNumber, pm.number
        attr_columns.append(None)  # col 1 = productId
        attr_columns.append(None)  # col 2 = serialNumber
        attr_columns.append(None)  # col 3 = pm.number
        attr_start = 4

    for ci in range(attr_start, max_col + 1):
        cell = ws.cell(1, ci)
        comment = cell.comment
        name, norm_type, lov_name = _parse_header_cell(
            cell.value, comment.text if comment else None, ci, result.errors
        )
        if name is not None and norm_type is not None:
            # 检查类型合法性
            if norm_type not in _VALID_TYPES:
                result.errors.append(
                    f"ATTRIBUTE_TYPE_NOT_FOUND: column {ci}: "
                    f"unknown type '{norm_type}' for attribute '{name}'"
                )
            attr_columns.append({
                "name": name,
                "type": norm_type,
                "lov_name": lov_name,
            })
        elif name is not None and norm_type is None:
            # 类型未识别（已在 _parse_header_cell 中添加错误）
            # 仍然记录列名以便后续处理（best-effort）
            attr_columns.append({
                "name": name,
                "type": "TEXT",
                "lov_name": None,
            })
        else:
            attr_columns.append(None)

    # ========== 重复属性校验 ==========
    seen: set[tuple[str, str]] = set()
    for col_def in attr_columns:
        if col_def is None:
            continue
        key = (col_def["name"], col_def["type"])
        if key in seen:
            result.errors.append(
                f"DUPLICATE_ATTRIBUTE: attribute '{col_def['name']}' "
                f"with type '{col_def['type']}' appears more than once"
            )
        seen.add(key)

    # ========== 解析数据行 ==========
    for row in range(2, ws.max_row + 1):
        if import_type == "pathdata":
            # pathdata: 零件号在第3列，停止条件检查第3列
            first_cell = ws.cell(row, 3)
        else:
            first_cell = ws.cell(row, 1)
        first_val = first_cell.value

        # Java 逻辑：cells[j][0]==null → break
        if first_val is None:
            # 检查其他列是否有值（仅 parts 导入）
            if import_type == "parts":
                other_has_value = False
                for ci in range(2, max_col + 1):
                    cv = ws.cell(row, ci).value
                    if cv is not None:
                        other_has_value = True
                        break
                if other_has_value:
                    result.errors.append(
                        f"EMPTY_FIELD: row {row}: part number is empty "
                        f"but other columns have values"
                    )
            break  # 停止解析后续行

        part_number = str(first_val).strip()
        part = ParsedPart(number=part_number)

        if import_type == "pathdata":
            # 提取 product_id 和 serial_number
            b_val = ws.cell(row, 1).value
            c_val = ws.cell(row, 2).value
            part.product_id = str(b_val).strip() if b_val is not None else None
            part.serial_number = str(c_val).strip() if c_val is not None else None

        # 解析各属性列的值
        for ci, col_def in enumerate(attr_columns, start=1):
            if col_def is None:
                continue
            cell = ws.cell(row, ci)
            cell_val = cell.value
            cell_comment = cell.comment
            attrs = _parse_data_cell(
                cell_val,
                cell_comment.text if cell_comment else None,
                col_def,
                row,
                result.errors,
            )
            part.attributes.extend(attrs)

        result.parts.append(part)

    return result
