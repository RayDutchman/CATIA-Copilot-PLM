"""Task 5: Import REST 端点测试（端到端）。"""
import io
import uuid
from openpyxl import Workbook
from openpyxl.comments import Comment
from fastapi.testclient import TestClient
from app.main import app

PREFIX = "/docdoku-plm-server-rest/api"
WS = "GD50"
client = TestClient(app)


def _token():
    r = client.post(f"{PREFIX}/auth/login", json={"login": "test1", "password": "password"})
    return r.headers.get("jwt")


def _make_test_xlsx(part_number="NOPE-IMPORT-TEST", attr_value="hello"):
    """合成最小合法 parts .xlsx（A1 comment=pm.number，B1=属性头，row2=数据）。"""
    wb = Workbook()
    ws = wb.worksheets[0]
    a1 = ws.cell(row=1, column=1, value="pm.number")
    a1.comment = Comment("pm.number", "importer")
    b1 = ws.cell(row=1, column=2, value="MyText <Text>")
    a2 = ws.cell(row=2, column=1, value=part_number)
    b2 = ws.cell(row=2, column=2, value=attr_value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _unique_filename():
    return f"test-import-{uuid.uuid4().hex}.xlsx"


# ── 测试 1: 预览端点 200 + 返回结构 ──────────────────────────────

def test_import_preview_attributes_returns_200_and_shape():
    h = {"Authorization": f"Bearer {_token()}"}
    data = _make_test_xlsx()
    url = f"{PREFIX}/workspaces/{WS}/parts/importPreview"
    r = client.post(
        url,
        files={"upload": (_unique_filename(), data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"importType": "attributes", "autoCheckout": "true"},
        headers=h,
    )
    assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
    body = r.json()
    assert "partRevsToCheckout" in body, f"Missing partRevsToCheckout: {body}"
    assert isinstance(body["partRevsToCheckout"], list)
    assert "partsToCreate" in body, f"Missing partsToCreate: {body}"
    assert isinstance(body["partsToCreate"], list)
    # 不存在的零件不应出现在 checkout 列表中
    assert body["partRevsToCheckout"] == []


# ── 测试 2: 非法 importType → 400 ───────────────────────────────

def test_import_invalid_type_400():
    h = {"Authorization": f"Bearer {_token()}"}
    data = _make_test_xlsx()
    url = f"{PREFIX}/workspaces/{WS}/parts/import"
    r = client.post(
        url,
        files={"upload": (_unique_filename(), data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"importType": "nonsense"},
        headers=h,
    )
    assert r.status_code == 400


# ── 测试 3: 属性导入不存在的零件 → 204 + 记录含错误 ──────────────

def test_import_attributes_nonexistent_part_204_and_record():
    h = {"Authorization": f"Bearer {_token()}"}
    filename = _unique_filename()
    data = _make_test_xlsx()
    url = f"{PREFIX}/workspaces/{WS}/parts/import"

    # 1) POST import → 204
    r = client.post(
        url,
        files={"upload": (filename, data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"importType": "attributes"},
        headers=h,
    )
    assert r.status_code == 204

    # 2) GET imports/{filename} → 200, 数组含 pending==false, succeed==false, 有 errors
    r2 = client.get(
        f"{PREFIX}/workspaces/{WS}/parts/imports/{filename}",
        headers=h,
    )
    assert r2.status_code == 200
    records = r2.json()
    assert isinstance(records, list)
    assert len(records) >= 1, f"No import record found for {filename}"

    record = records[0]
    import_id = record["id"]
    assert record["pending"] is False
    assert record["succeed"] is False
    assert len(record["errors"]) > 0
    assert any("PartMasterNotFound" in e for e in record["errors"]), \
        f"Expected PartMasterNotFound in errors: {record['errors']}"

    # 3) DELETE import/{id} → 204
    r3 = client.delete(
        f"{PREFIX}/workspaces/{WS}/parts/import/{import_id}",
        headers=h,
    )
    assert r3.status_code == 204

    # 4) GET imports/{filename} again → 该 id 不再出现
    r4 = client.get(
        f"{PREFIX}/workspaces/{WS}/parts/imports/{filename}",
        headers=h,
    )
    assert r4.status_code == 200
    remaining = [rec for rec in r4.json() if rec["id"] == import_id]
    assert len(remaining) == 0, f"Import {import_id} was not deleted"


# ── 测试 4: BOM 类型导入 → 204 + 记录 failed ────────────────────

def test_import_bom_type_returns_204_record_failed():
    h = {"Authorization": f"Bearer {_token()}"}
    filename = _unique_filename()
    data = _make_test_xlsx()
    url = f"{PREFIX}/workspaces/{WS}/parts/import"

    # 1) POST import → 204
    r = client.post(
        url,
        files={"upload": (filename, data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"importType": "bom"},
        headers=h,
    )
    assert r.status_code == 204

    # 2) GET → succeed false
    r2 = client.get(
        f"{PREFIX}/workspaces/{WS}/parts/imports/{filename}",
        headers=h,
    )
    assert r2.status_code == 200
    records = r2.json()
    assert len(records) >= 1
    record = records[0]
    assert record["succeed"] is False

    # 3) Cleanup
    client.delete(
        f"{PREFIX}/workspaces/{WS}/parts/import/{record['id']}",
        headers=h,
    )


# ── 测试 5: GET 不存在的 import id → 404 ────────────────────────

def test_get_single_import_404():
    h = {"Authorization": f"Bearer {_token()}"}
    r = client.get(
        f"{PREFIX}/workspaces/{WS}/parts/import/does-not-exist",
        headers=h,
    )
    assert r.status_code == 404
