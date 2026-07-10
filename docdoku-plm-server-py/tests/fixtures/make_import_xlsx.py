"""合成 .xlsx 夹具生成器，用于 Excel 导入解析器测试。

用 openpyxl 写带 cell-comment 的 .xlsx，返回 bytes。
"""

import io
import datetime
from openpyxl import Workbook
from openpyxl.comments import Comment


_AUTHOR = "importer"


def make_xlsx(columns: list[dict], data_rows: list[list[str | None] | dict]) -> bytes:
    """通用 xlsx 构建器。

    columns: [{"header": "文本值", "comment": "comment文本" | None}, ...]
    第1列为 pm.number 列，其 comment 必须是 "pm.number"（parts 导入）。
    data_rows: 每行可以是 list[str | None]（纯值）或
               {"values": [...], "comments": [...]}（带批注的数据单元格）。
    """
    wb = Workbook()
    ws = wb.worksheets[0]

    # 写表头
    for ci, col_spec in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_spec["header"])
        comment_text = col_spec.get("comment")
        if comment_text is not None:
            cell.comment = Comment(comment_text, _AUTHOR)

    # 写数据行
    for ri, row in enumerate(data_rows, start=2):
        if isinstance(row, dict):
            values = row.get("values", [])
            comments = row.get("comments", [])
        else:
            values = row
            comments = []

        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            # 写该单元格的 comment
            if ci - 1 < len(comments):
                c = comments[ci - 1]
                if c is not None:
                    cell.comment = Comment(str(c), _AUTHOR)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_valid_parts_xlsx() -> bytes:
    """合法 parts 导入文件：
    - A1 comment "pm.number"
    - 多个属性列（Text/Number/Boolean/Date/URL/Long_Text/LOV）
    - 数据行含正确值
    - 部分数据单元格带数字 comment(属性ID)模拟"更新"，部分无 comment 模拟"新建"
    """
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},

        {"header": "Color <ListOfValues> <Colors_LOV>", "comment": None},
        {"header": "Weight <Number>", "comment": None},
        {"header": "Description <Text>", "comment": None},
        {"header": "Active <Boolean>", "comment": None},
        {"header": "ReleaseDate <Date>", "comment": None},
        {"header": "DocURL <URL>", "comment": None},
        {"header": "Notes <Long_Text>", "comment": None},
    ]

    # line-data: 更新已有属性（带 comment 含属性ID）
    # 结构: [{"header": "PartNumber", "comment": "pm.number"}, ...]
    # 每个数据行: dict with values, comments
    # 共 7 个属性列 + 1 个零件号列 = 8 列
    data_rows: list[dict] = []

    # 第2行：带属性 ID 的更新行（模拟更新已有属性）
    data_rows.append({
        "values": [
            "PART-001",        # 零件号
            "Red",             # Color (LOV)
            "2.5",             # Weight (Number)
            "Some text",       # Description (Text)
            "true",            # Active (Boolean)
            "2025-01-15 00:00:00",  # ReleaseDate (Date)
            "https://example.com/doc",  # DocURL (URL)
            "Long description text",   # Notes (Long_Text)
        ],
        "comments": [
            None,     # PartNumber 列无 comment
            "101",    # Color = 属性ID 101 (更新)
            "102",    # Weight = 属性ID 102 (更新)
            "103",    # Description = 属性ID 103 (更新)
            "104",    # Active = 属性ID 104 (更新)
            "105",    # ReleaseDate = 属性ID 105 (更新)
            "106",    # DocURL = 属性ID 106 (更新)
            "107",    # Notes = 属性ID 107 (更新)
        ],
    })

    # 第3行：无 comment（模拟新建属性）
    data_rows.append({
        "values": [
            "PART-002",
            "Blue",
            "3.14",
            "Another text",
            "false",
            "2025-02-01 00:00:00",
            "ftp://files.example.com/doc",
            "",  # Long_Text 允许空值
        ],
        "comments": [
            None, None, None, None, None, None, None, None,
        ],
    })

    # 第4行：部分有 comment 部分无（混合）
    data_rows.append({
        "values": [
            "PART-003",
            "Green",
            "0.0",
            "Mixed attributes",
            "true",
            "2025-03-10 00:00:00",
            "http://example.com",
            "Some notes",
        ],
        "comments": [
            None,
            None,    # LOV 新建
            "202",   # Number 更新
            None,    # Text 新建
            "204",   # Boolean 更新
            None,    # Date 新建
            "206",   # URL 更新
            None,    # Long_Text 新建
        ],
    })

    return make_xlsx(columns, data_rows)


def make_invalid_date_parts_xlsx() -> bytes:
    """DATE 列含非法日期值。"""
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "MyDate <Date>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", "not-a-date"], "comments": [None, "101"]},
    ]
    return make_xlsx(columns, data_rows)


def make_missing_pmnumber_comment_xlsx() -> bytes:
    """A1 缺少 pm.number comment。"""
    columns = [
        {"header": "PartNumber", "comment": None},  # 无 pm.number comment
        {"header": "Attr <Text>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", "value"], "comments": [None, "101"]},
    ]
    return make_xlsx(columns, data_rows)


def make_empty_xlsx() -> bytes:
    """空 workbook（无 sheet / 全是空行）。"""
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_text_too_long_xlsx() -> bytes:
    """TEXT 列值超过 255 字符。"""
    long_text = "A" * 300
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "Desc <Text>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", long_text], "comments": [None, None]},
    ]
    return make_xlsx(columns, data_rows)


def make_invalid_boolean_xlsx() -> bytes:
    """BOOLEAN 列值不是 true/false。"""
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "Active <Boolean>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", "yes"], "comments": [None, "101"]},
    ]
    return make_xlsx(columns, data_rows)


def make_duplicate_attr_xlsx() -> bytes:
    """两列表头属性名和类型相同。"""
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "Desc <Text>", "comment": None},
        {"header": "Desc <Text>", "comment": None},  # 重复
    ]
    data_rows = [
        {"values": ["PART-001", "val1", "val2"], "comments": [None, None, None]},
    ]
    return make_xlsx(columns, data_rows)


def make_multi_value_xlsx() -> bytes:
    """多值单元格：值 "a|b"、comment "12|34"。"""
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "Tags <Text>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", "a|b"], "comments": [None, "12|34"]},
    ]
    return make_xlsx(columns, data_rows)


def make_multi_value_more_values_than_ids_xlsx() -> bytes:
    """多值单元格：值比 ID 多。"""
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "Tags <Text>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", "a|b|c"], "comments": [None, "12"]},
    ]
    return make_xlsx(columns, data_rows)


def make_empty_first_col_stop_xlsx() -> bytes:
    """数据中间某行第一列为空，之后行应停止解析。"""
    columns = [
        {"header": "PartNumber", "comment": "pm.number"},
        {"header": "Attr <Text>", "comment": None},
    ]
    data_rows = [
        {"values": ["PART-001", "val1"], "comments": [None, None]},
        {"values": [None, "val2"], "comments": [None, None]},       # 首列空 → 停止
        {"values": ["PART-003", "val3"], "comments": [None, None]},  # 应被忽略
    ]
    return make_xlsx(columns, data_rows)


def make_pathdata_xlsx() -> bytes:
    """pathdata 导入文件。
    - A1 comment "ctx.productId"
    - B1 comment "ctx.serialNumber"
    - C1 comment "pm.number"
    - 后面跟属性列（至少一个，让总列数 > 3）
    """
    columns = [
        {"header": "ProductX", "comment": "ctx.productId"},
        {"header": "SN001", "comment": "ctx.serialNumber"},
        {"header": "PartNum", "comment": "pm.number"},
        {"header": "Weight <Number>", "comment": None},
    ]
    data_rows = [
        {
            "values": ["PROD-001", "SN-001", "PART-001", "1.5"],
            "comments": [None, None, None, "201"],
        },
        {
            "values": ["PROD-002", "SN-002", "PART-002", "2.0"],
            "comments": [None, None, None, None],
        },
    ]
    return make_xlsx(columns, data_rows)
